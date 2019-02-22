import openpyxl
import os
import sys

# data_only=True, for read value, otherwise read the math equation in cell
fgo_xls  = openpyxl.load_workbook('FGO.xlsx', data_only=True)
fgo_cmd  = open('fgo_cmd.js', 'w+')
fgo_func = open('fgo_func.js', 'r', encoding='UTF-8')
# print(type(fgo_xls))

worksheets = fgo_xls.sheetnames
print(worksheets)

# sheet = fgo_xls.get_sheet_by_name('Pexel LUT')
sheet = fgo_xls['Pexel LUT']
# print(sheet)
# print(sheet.title)

# print(fgo_xls.active)
# print(sheet['B1'], sheet['B1'].value)
# print(sheet.cell(row=1, column=2).value)
# print(sheet.max_row, sheet.max_column)

basic_func_name = []  # column=1
basic_func_calx = []  # column=4
basic_func_caly = []  # column=5
basic_func_slpt = []  # column=6
basic_func_skil = []  # column=8

# from row 3 to end
row_start = 3
row_end   = sheet.max_row + 1

print(row_start, row_end)

for i in range(row_start, row_end):
	basic_func_name.append(sheet.cell(row=i, column=1).value)

for i in range(row_start, row_end):
	basic_func_calx.append(sheet.cell(row=i, column=4).value)

for i in range(row_start, row_end):
	basic_func_caly.append(sheet.cell(row=i, column=5).value)

for i in range(row_start, row_end):
	basic_func_slpt.append(sheet.cell(row=i, column=6).value)

for i in range(row_start, row_end):
	basic_func_skil.append(sheet.cell(row=i, column=8).value)

print('\n\n')
print('+-----+------------+-----------+------------+---+')
print('| No  | KEY        | POSITION  | DURATION   | F |')
print('+-----+------------+-----------+------------+---+')
for i in range(len(basic_func_name)):
	print('+-----+------------+-----------+------------+---+')
	print('| %3d | %-10s | %4d,%4d | %7d ms | %1d |' % (i, basic_func_name[i], basic_func_calx[i], basic_func_caly[i], basic_func_slpt[i], basic_func_skil[i]))
	
	if i==len(basic_func_name)-1:
		print('+-----+------------+-----------+------------+---+')

# generate fgo_cmd.js file

fgo_cmd.write('var CMD = {};\n\n')

fgo_cmd.write(fgo_func.read())
fgo_cmd.write('\n\n')

for i in range(len(basic_func_name)):

	if basic_func_skil[i]:
		fgo_cmd.write('CMD.%s = function(all, sel, duration){\n' % basic_func_name[i])
		# fgo_cmd.write('function %s(all, sel, duration){\n' % basic_func_name[i])
		fgo_cmd.write('    // \'all\' and \'duration\' are multi-use ...\n')
		fgo_cmd.write('    // when only para \'all\', use it as duration ...\n')
		fgo_cmd.write('    // ex : A1 skill to \'sel\' / \'all\' in sequence ...\n\n')
		fgo_cmd.write('    var all      = arguments[0] ? arguments[0] : 0;\n')
		fgo_cmd.write('    var sel      = arguments[1] ? arguments[1] : 0;\n')
		fgo_cmd.write('    var duration = arguments[0] ? arguments[0] : %d;\n\n' % basic_func_slpt[i])
		fgo_cmd.write('    // if it is a specific skill for someone\n')
		fgo_cmd.write('    // then use 1.5ms instead 3s when no duration parameter\n')
		fgo_cmd.write('    if ( all * sel != 0 )\n')
		fgo_cmd.write('        duration = arguments[2] ? arguments[2] : 1500;\n\n')	
		fgo_cmd.write('    click(%s, %s);\n' % (str(basic_func_calx[i]), str(basic_func_caly[i])))
		fgo_cmd.write('    sleep(duration);\n\n')
		fgo_cmd.write('    if ( all * sel != 0 )\n')
		fgo_cmd.write('        CMD.SKILL_SELXX(all, sel);\n')
		fgo_cmd.write('};\n\n')
	else:
		fgo_cmd.write('CMD.%s = function(duration){\n' % basic_func_name[i])
		# fgo_cmd.write('function %s(duration){\n' % basic_func_name[i])
		fgo_cmd.write('    var duration = arguments[0] ? arguments[0] : %d;\n' % basic_func_slpt[i])
		fgo_cmd.write('    click(%s, %s);\n' % (str(basic_func_calx[i]), str(basic_func_caly[i])))
		fgo_cmd.write('    sleep(duration);\n')
		fgo_cmd.write('};\n\n')

fgo_cmd.write('module.exports = CMD;\n\n')




fgo_cmd.close()
























